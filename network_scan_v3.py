import psutil
import time
import socket
import requests
import openpyxl
import ipaddress
from openpyxl.styles import PatternFill, Font, Border, Side

VT_API_KEY = "YOUR_VT_API_KEY"
ABUSEIPDB_API_KEY = "YOUR_ABUSEIPDB_API_KEY"

SCAN_DURATION = 100
OUTPUT_FILE = "network_results_full_analytics.xlsx"


# ==========================
# SERVICE DETECTION
# ==========================
def get_service(port):
    try:
        return socket.getservbyport(port)
    except:
        return "Unknown"


# ==========================
# HELPERS
# ==========================
def is_private_ip(ip):
    try:
        return ipaddress.ip_address(ip).is_private
    except:
        return True

def reverse_dns(ip):
    try:
        return socket.gethostbyaddr(ip)[0]
    except:
        return ""

def ipinfo_lookup(ip):
    try:
        r = requests.get(f"https://ipinfo.io/{ip}/json", timeout=5)
        if r.status_code == 200:
            data = r.json()
            country = data.get("country", "")
            org = data.get("org", "")
            return country, org
    except:
        pass
    return "", ""

def vt_score(ip):
    try:
        url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
        headers = {"x-apikey": VT_API_KEY}
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code != 200:
            return None
        data = r.json()
        stats = data.get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
        return stats.get("malicious", 0)
    except:
        return None

def abuseipdb_score(ip):
    try:
        url = "https://api.abuseipdb.com/api/v2/check"
        r = requests.get(url, params={"ipAddress": ip, "maxAgeInDays": 90},
                         headers={"Key": ABUSEIPDB_API_KEY}, timeout=10)
        if r.status_code != 200:
            return None
        data = r.json()
        return data.get("data", {}).get("abuseConfidenceScore", 0)
    except:
        return None

def risk_color(vt, abuseipdb):
    if vt is None and abuseipdb is None:
        return "FFFFFF"
    risk = 0
    risk += (vt or 0) * 5
    risk += (abuseipdb or 0)
    if risk > 60:
        return "FF0000"
    if risk > 20:
        return "FFFF00"
    return "00FF00"


# ==========================
# SCANNING
# ==========================
print("-> Monitoring the network for 100 seconds...")
start_time = time.time()
found = {}

while time.time() - start_time < SCAN_DURATION:
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            for conn in proc.net_connections(kind='inet'):
                if conn.raddr and conn.status == 'ESTABLISHED':
                    ip = conn.raddr.ip
                    port = conn.raddr.port
                    if not is_private_ip(ip):

                        # Accessed files by the process (optional / may require permissions)
                        files = []
                        try:
                            files = [f.path for f in proc.open_files()]
                        except Exception:
                            pass

                        found[(ip, port, proc.info['pid'])] = {
                            "process": proc.info['name'],
                            "pid": proc.info['pid'],
                            "files": files
                        }

        except:
            pass
    time.sleep(1)

print(f"🔍 Found {len(found)} unique external connections. Beginning scoring...")


# ==========================
# EXCEL OUTPUT
# ==========================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Network Scan"

ws.append([
    "IP", "Domain", "Port", "Service", "Process", "PID", "Country", "AS Name",
    "VT Score", "AbuseIPDB Score", "Wireshark Filter", "Accessed Files"
])

header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin_gray = Side(border_style="thin", color="808080")
header_border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = header_border


for (ip, port, pid), info in found.items():
    dom = reverse_dns(ip)
    country, asname = ipinfo_lookup(ip)
    vt = vt_score(ip)
    abuseipdb = abuseipdb_score(ip)
    service = get_service(port)
    wireshark_filter = f"ip.addr=={ip} && tcp.port=={port}"
    accessed_files = "; ".join(info['files']) if info['files'] else ""

    ws.append([
        ip,
        dom,
        port,
        service,
        info["process"],
        info["pid"],
        country,
        asname,
        vt,
        abuseipdb,
        wireshark_filter,
        accessed_files
    ])


# Add color risk visualization
for row in ws.iter_rows(min_row=2):
    vt_cell = row[8]
    abuseipdb_cell = row[9]

    vt_color = risk_color(vt_cell.value, 0)
    vt_cell.fill = PatternFill(start_color=vt_color, end_color=vt_color, fill_type="solid")

    abuseipdb_color = risk_color(0, abuseipdb_cell.value)
    abuseipdb_cell.fill = PatternFill(start_color=abuseipdb_color, end_color=abuseipdb_color, fill_type="solid")


wb.save(OUTPUT_FILE)
print(f"✅ Completed. Excel file generated: {OUTPUT_FILE}")
